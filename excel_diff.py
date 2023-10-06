import openpyxl
from openpyxl.styles import PatternFill, Font

# 変化点を検出して出力先のエクセルに色付ける関数
def highlight_differences(sheet, row1, row2, output_row, highlight_color="52C58C"):
    for i, (cell1, cell2) in enumerate(zip(row1, row2)):
        if cell1 != cell2:
            output_cell = sheet.cell(row=output_row, column=i + 1)
            output_cell.value = cell2
            output_cell.fill = PatternFill(start_color=highlight_color, end_color=highlight_color, fill_type="solid")
            output_cell.font = Font(color="FF0000")  # 赤文字に設定
        else:
            output_cell = sheet.cell(row=output_row, column=i + 1)
            output_cell.value = cell2         

# メイン処理を実行する関数
def main():
    file1 = 'in/test1.xlsx'
    file2 = 'in/test2.xlsx'
    result = 'out/result.xlsx'
    # 2つのエクセルファイルを読み込む
    workbook1 = openpyxl.load_workbook(file1)
    workbook2 = openpyxl.load_workbook(file2)
    
    # シートを選択（必要に応じてシート名を変更）
    sheet1 = workbook1['Sheet1']
    sheet2 = workbook2['Sheet1']
    
    # データをバッファに格納
    data1 = [row for row in sheet1.iter_rows(values_only=True)]
    data2 = [row for row in sheet2.iter_rows(values_only=True)]
    
    # 出力先のエクセルを作成
    output_workbook = workbook2
    output_sheet = output_workbook.active
    
    # file2で挿入されている行を確認し、file1に空白の行を挿入する
    for row_num, (row_data1, row_data2) in enumerate(zip(data1, data2), start=1):
        found = False
        # 挿入されている行を確認
        if row_data2[0] == None:
            found = True
        # file2で挿入されている行と同一の行に空白行を挿入する
        if found:
            data1.insert(row_num - 1, [''] * len(row_data2))
        # 最後まで確認終えたら終了
        if row_data2[0] == 'end':
            break

    # 変化点を検出して出力先のエクセルに書き込み
    for row_num, (row_data1, row_data2) in enumerate(zip(data1, data2), start=1):
        if row_data1 != row_data2:
            highlight_differences(output_sheet, row_data1, row_data2, row_num)
    
    # 出力先のエクセルを保存
    output_workbook.save(result)

if __name__ == "__main__":
    main()
